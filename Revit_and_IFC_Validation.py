#!/usr/bin/env python
# coding: utf-8

# In[2]:


##################################################################################################

file_path_excel = r'C:\Users\Artem\Desktop\DDC\Validation\DDC Revit and IFC Validation.xlsx'

##################################################################################################

# If you do not have pandas and openpyxl installed, install the libraries
# pip install pandas openpyxl

import os, subprocess, pandas as pd, datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

# Setting up the paths and other constants
path_conv, file_path = pd.read_excel(file_path_excel).iloc[[2, 0], 2]
output_file = file_path[:-4] + "_ifc.xlsx" if file_path.lower().endswith(".ifc") else file_path[:-4] + "_rvt.xlsx"
output_timestamped = file_path_excel[:-5] + f"_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# Determine the correct exporter and data column based on the file extension
exporter = "IfcExporter.exe" if file_path.lower().endswith(".ifc") else "RvtExporter.exe"

# Running the conversion process and reading the resulting data
subprocess.run([os.path.join(path_conv, exporter), file_path], cwd=path_conv, check=True)
df = pd.read_excel(output_file)

# Clean column names if they contain ':'
if any(':' in col for col in df.columns):
    df.columns = [col.split(' : ')[0] for col in df.columns]

# Loading the workbook and setting up the worksheet for data update
workbook = load_workbook(file_path_excel)
sheet = workbook.active
df_excel = pd.read_excel(file_path_excel)
ifc_classes = df_excel['Unnamed: 1'][5:]
properties_parameters = df_excel['Unnamed: 2'][5:]

# Define color fills
red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')

# Updating data in the Excel file
for idx, (ifc_class, prop_param) in enumerate(zip(ifc_classes, properties_parameters), start=6):
    if prop_param in df.columns:
        filtered_df = df[df["Category"].eq(ifc_class) & df[prop_param].notna()]
        fill = filtered_df[prop_param].notna().mean()  # Calculate the fill percentage
        unique_vals = filtered_df[prop_param].nunique()  # Count unique values
        unique_list = ", ".join(filtered_df[prop_param].dropna().unique().astype(str))  # List unique values
    else:
        fill = 0  # Set fill percentage to 0 if column is missing
        unique_vals = 0  # Set unique values count to 0
        unique_list = ""

    # Set values and formats for Excel cells
    cell = sheet[f'D{idx+1}']
    cell.value, cell.number_format, cell.alignment = fill, '0.00%', Alignment(horizontal='center')
    cell.fill = green_fill if fill > 0 else red_fill
    sheet[f'E{idx+1}'].value, sheet[f'E{idx+1}'].alignment = unique_vals, Alignment(horizontal='center')
    sheet[f'F{idx+1}'].value, sheet[f'F{idx+1}'].alignment = unique_list, Alignment(horizontal='left')

# Saving the changes back to a new timestamped Excel file
workbook.save(output_timestamped)

# Outputting a completion message with the name of the file
print(f"Report is ready with the name: {output_timestamped}")


# In[ ]:




